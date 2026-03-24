from __future__ import annotations

import io
import logging
import os
import subprocess
import tempfile

try:
    import pdfplumber  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover
    pdfplumber = None  # type: ignore[assignment]

try:
    import docx  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover
    docx = None  # type: ignore[assignment]

try:
    import openpyxl  # type: ignore[import-untyped]
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore[assignment]

LOGGER = logging.getLogger(__name__)

ATTACHMENT_INLINE_TOKEN_BUDGET: int = int(os.environ.get("ATTACHMENT_INLINE_TOKEN_BUDGET", "4000"))

_COMPLEX_DOCX_TEXT_THRESHOLD = 200

_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_CSV_MIME = "text/csv"
_PDF_MIME = "application/pdf"


def _estimate_tokens(text: str) -> int:
    return len(text) // 4


def _convert_docx(data: bytes) -> str:
    if docx is None:
        return "[attachment: python-docx is not installed; docx conversion unavailable]"
    try:
        doc = docx.Document(io.BytesIO(data))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
        if len(text.strip()) < _COMPLEX_DOCX_TEXT_THRESHOLD:
            hint = (
                "For best results with text + images, send a .md or .txt file"
                " and attach images separately."
            )
            text = f"{text}\n\n{hint}".strip()
        return text
    except Exception as exc:  # noqa: BLE001
        return f"[attachment: docx conversion failed — {exc}]"


def _convert_xlsx(data: bytes) -> str:
    if openpyxl is None:
        return "[attachment: openpyxl is not installed; xlsx conversion unavailable]"
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_parts: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        raw_rows = list(ws.iter_rows())
        if not raw_rows:
            continue
        rows = [
            [str(getattr(cell, "value", cell) if getattr(cell, "value", cell) is not None else "") for cell in row]
            for row in raw_rows
        ]
        col_widths = [max(len(r[c]) for r in rows) for c in range(len(rows[0]))]
        lines: list[str] = []
        for ridx, row in enumerate(rows):
            padded = " | ".join(cell.ljust(col_widths[c]) for c, cell in enumerate(row))
            lines.append(f"| {padded} |")
            if ridx == 0:
                sep = " | ".join("-" * col_widths[c] for c in range(len(rows[0])))
                lines.append(f"| {sep} |")
        sheet_parts.append(f"**Sheet: {sheet_name}**\n\n" + "\n".join(lines))
    return "\n\n".join(sheet_parts)


def _convert_csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="replace")
    return "\n".join(line.rstrip("\r") for line in text.splitlines())


def _convert_pdf(data: bytes) -> str:
    if pdfplumber is None:
        return "[attachment: pdfplumber is not installed; pdf conversion unavailable]"
    try:
        pages: list[str] = []
        with pdfplumber.open(io.BytesIO(data)) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    pages.append(page_text)
        return "\n\n".join(pages)
    except Exception as exc:  # noqa: BLE001
        return f"[attachment: pdf conversion failed — {exc}]"


def convert_attachment(filename: str, mime_type: str, data: bytes) -> str:
    """Convert an attachment to extracted text.

    Returns the extracted text, or a fallback notice when the type is
    unsupported or a required library is missing.
    """
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    mime = (mime_type or "").split(";")[0].strip().lower()

    if ext == ".docx" or mime == _DOCX_MIME:
        return _convert_docx(data)
    if ext == ".xlsx" or mime == _XLSX_MIME:
        return _convert_xlsx(data)
    if ext == ".csv" or mime == _CSV_MIME or mime.startswith("text/csv"):
        return _convert_csv(data)
    if ext == ".pdf" or mime == _PDF_MIME:
        return _convert_pdf(data)

    return f"[attachment: {filename} — unsupported type {mime_type}, skipped]"


def stage_attachment_to_container(
    *,
    filename: str,
    data: bytes,
    container_name: str,
    workdir: str = "/workspace/repo",
) -> str:
    """Write data to a temp file and copy it into the container via podman cp.

    Returns the in-container path.
    """
    import os as _os

    ext = ("." + filename.rsplit(".", 1)[-1]) if "." in filename else ".bin"
    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
        tmp.write(data)
        tmp_path = tmp.name
    try:
        container_dir = "/tmp/uploads"
        subprocess.run(
            ["podman", "exec", container_name, "mkdir", "-p", container_dir],
            check=False,
            capture_output=True,
            text=True,
        )
        dest = f"{container_name}:{container_dir}/{filename}"
        subprocess.run(
            ["podman", "cp", tmp_path, dest],
            check=True,
            capture_output=True,
            text=True,
        )
    finally:
        if _os.path.exists(tmp_path):
            _os.unlink(tmp_path)
    return f"{container_dir}/{filename}"


def attachment_to_prompt_fragment(
    *,
    filename: str,
    mime_type: str,
    data: bytes,
    container_name: str | None = None,
    workdir: str = "/workspace/repo",
) -> str:
    """Convert an attachment and return a prompt-ready string.

    If the extracted text fits within ATTACHMENT_INLINE_TOKEN_BUDGET, it is
    returned inline.  If it is over budget and a container_name is provided,
    the file is staged and a Read-tool pointer is returned instead.
    """
    text = convert_attachment(filename, mime_type, data)
    if _estimate_tokens(text) <= ATTACHMENT_INLINE_TOKEN_BUDGET:
        return f"[attachment: {filename}]\n{text}"

    if container_name:
        try:
            container_path = stage_attachment_to_container(
                filename=filename,
                data=data,
                container_name=container_name,
                workdir=workdir,
            )
            return (
                f"[attachment: {filename} has been placed at {container_path}."
                f" Use your Read tool to examine it.]"
            )
        except Exception as exc:  # noqa: BLE001
            LOGGER.warning(
                "file_converter.stage_failed filename=%s error=%s", filename, exc
            )

    return f"[attachment: {filename}]\n{text}"
